"""
excel_exporter.py — Pandas + Openpyxl ile renkli Excel çıktısı.

Tasarım
-------
- Başlık satırı: koyu mavi arka plan, beyaz kalın yazı
- Geçti satırı: açık yeşil (#C8E6C9)
- Kaldı satırı : açık kırmızı (#FFCDD2)
- Belirsiz      : açık sarı  (#FFF9C4)
- Otomatik sütun genişliği
- Sayı formatı: 2 ondalık basamak
"""
from __future__ import annotations
import io
import logging
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session
from app.models import Student, Grade, Class

logger = logging.getLogger(__name__)

# Renk paletleri
COLOR_HEADER_BG = "1565C0"   # Koyu mavi
COLOR_HEADER_FG = "FFFFFF"   # Beyaz
COLOR_GECTI     = "C8E6C9"   # Açık yeşil
COLOR_KALDI     = "FFCDD2"   # Açık kırmızı
COLOR_BELIRSIZ  = "FFF9C4"   # Açık sarı
COLOR_ALT_ROW   = "F5F5F5"   # Zebra şeridi


def _row_color(status: str) -> str:
    match status:
        case "Geçti":    return COLOR_GECTI
        case "Kaldı":    return COLOR_KALDI
        case _:          return COLOR_BELIRSIZ


def build_excel(db: Session, class_id: int) -> bytes:
    """
    class_id'ye göre tüm öğrenci notlarını çek ve renkli Excel üret.
    Döndürülen bytes doğrudan HTTP response olarak gönderilebilir.
    """
    cls: Class | None = db.query(Class).filter(Class.id == class_id).first()
    if cls is None:
        raise ValueError(f"Sınıf bulunamadı: id={class_id}")

    rows: list[dict[str, Any]] = []
    for student in cls.students:
        # En güncel notu al
        grade: Grade | None = (
            db.query(Grade)
            .filter(Grade.student_id == student.id)
            .order_by(Grade.updated_at.desc())
            .first()
        )
        avg = grade.calculated_average if grade else None
        status = "Belirsiz"
        if avg is not None:
            status = "Geçti" if avg >= 50 else "Kaldı"

        rows.append({
            "Okul No":      student.school_no,
            "Adı Soyadı":   student.name,
            "1. Sınav":     grade.exam1 if grade else None,
            "2. Sınav":     grade.exam2 if grade else None,
            "1. Perf":      grade.perf1 if grade else None,
            "2. Perf":      grade.perf2 if grade else None,
            "Ortalama":     avg,
            "Durum":        status,
        })

    if not rows:
        logger.warning("Sınıf %d için kayıt bulunamadı.", class_id)
        rows.append({k: None for k in
                     ["Okul No", "Adı Soyadı", "1. Sınav", "2. Sınav",
                      "1. Perf", "2. Perf", "Ortalama", "Durum"]})

    df = pd.DataFrame(rows)
    df.insert(0, "#", range(1, len(df) + 1))

    # Pandas → Excel buffer
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name=cls.name,
            index=False,
            startrow=1,   # 1. satır başlık için ayrıldı
        )
        workbook  = writer.book
        worksheet = writer.sheets[cls.name]

        # ── Sınıf adı başlık satırı ──────────────────────────────────────────
        title_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title_font = Font(color=COLOR_HEADER_FG, bold=True, size=13)
        worksheet.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"{cls.name} — Not Dökümü"
        title_cell.fill  = title_fill
        title_cell.font  = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Sütun başlık satırı (satır 2) ────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1976D2")
        header_font = Font(color=COLOR_HEADER_FG, bold=True, size=11)
        thin_side   = Side(style="thin", color="BDBDBD")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        # ── Veri satırları ────────────────────────────────────────────────────
        status_col_idx = df.columns.get_loc("Durum") + 1  # 1-indexed

        for row_idx, row_data in enumerate(rows, start=3):  # satır 3'ten başlar
            status = row_data.get("Durum", "Belirsiz")
            row_fill = PatternFill("solid", fgColor=_row_color(status))

            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill   = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Sayı formatı
                col_name = df.columns[col_idx - 1]
                if col_name in ("1. Sınav", "2. Sınav", "1. Perf", "2. Perf", "Ortalama"):
                    cell.number_format = "0.00"

        # ── Otomatik sütun genişliği ──────────────────────────────────────────
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[len(str(r.get(col_name, "") or "")) for r in rows]
            )
            worksheet.column_dimensions[col_letter].width = min(max_len + 4, 30)

        # İlk satır yüksekliği
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 22

        # Satırları dondur (başlıklar sabit kalsın)
        worksheet.freeze_panes = "A3"

    buffer.seek(0)
    return buffer.getvalue()
