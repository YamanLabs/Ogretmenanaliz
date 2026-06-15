"""
excel_growth_exporter.py — Pandas + Openpyxl ile Gelişim Değerlendirme Excel çıktısı.
"""
from __future__ import annotations
import io
import logging
from typing import Any

import pandas as pd
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session
from app.models import Class, Grade

logger = logging.getLogger(__name__)

# Renk paletleri (E-okul laciverti)
COLOR_HEADER_BG = "2B3E50"   # Dominant lacivert
COLOR_HEADER_FG = "FFFFFF"   # Beyaz
COLOR_ALT_ROW   = "F8F9FA"   # Açık zebra şeridi
COLOR_WHITE     = "FFFFFF"


def build_growth_excel(db: Session, class_id: int) -> bytes:
    """
    class_id'ye göre tüm öğrencilerin gelişim notlarını çek ve renkli Excel üret.
    Döndürülen bytes doğrudan HTTP response olarak gönderilebilir.
    """
    cls: Class | None = db.query(Class).filter(Class.id == class_id).first()
    if cls is None:
        raise ValueError(f"Sınıf bulunamadı: id={class_id}")

    rows: list[dict[str, Any]] = []
    for student in cls.students:
        # En güncel notu al
        grade: Grade | None = (
            db.query(Grade)
            .filter(Grade.student_id == student.id)
            .order_by(Grade.updated_at.desc())
            .first()
        )

        rows.append({
            "Okul No":      student.school_no,
            "Adı Soyadı":   student.name,
            "Sınıf İçi Ders Katılım":   grade.growth_attendance if grade else None,
            "Sınıf Dışı Faaliyetler":   grade.growth_activities if grade else None,
            "Ürün Değerlendirme":      grade.growth_product if grade else None,
            "Sosyal Duygusal Akademik Gelişim": grade.growth_social_emotional if grade else None,
            "Öğrenci Gelişimi":        grade.growth_progress if grade else None,
        })

    if not rows:
        logger.warning("Sınıf %d için kayıt bulunamadı.", class_id)
        rows.append({k: None for k in
                     ["Okul No", "Adı Soyadı", "Sınıf İçi Ders Katılım", 
                      "Sınıf Dışı Faaliyetler", "Ürün Değerlendirme", 
                      "Sosyal Duygusal Akademik Gelişim", "Öğrenci Gelişimi"]})

    df = pd.DataFrame(rows)
    df.insert(0, "#", range(1, len(df) + 1))

    # Pandas → Excel buffer
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name="Gelisim Raporu",
            index=False,
            startrow=1,   # 1. satır başlık için ayrıldı
        )
        workbook  = writer.book
        worksheet = writer.sheets["Gelisim Raporu"]

        # ── Sınıf adı başlık satırı ──────────────────────────────────────────
        title_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        title_font = Font(color=COLOR_HEADER_FG, bold=True, size=13)
        worksheet.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"{cls.name} — Gelişim Değerlendirme Raporu"
        title_cell.fill  = title_fill
        title_cell.font  = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Sütun başlık satırı (satır 2) ────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="34495E") # Yardımcı koyu mavi/gri
        header_font = Font(color=COLOR_HEADER_FG, bold=True, size=11)
        thin_side   = Side(style="thin", color="BDBDBD")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        # ── Veri satırları ────────────────────────────────────────────────────
        for row_idx, row_data in enumerate(rows, start=3):  # satır 3'ten başlar
            # Alternatif satır renklendirme (Zebra şerit)
            bg_color = COLOR_ALT_ROW if row_idx % 2 == 0 else COLOR_WHITE
            row_fill = PatternFill("solid", fgColor=bg_color)

            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill   = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Sayı formatı
                col_name = df.columns[col_idx - 1]
                if col_name not in ("#", "Okul No", "Adı Soyadı"):
                    cell.number_format = "0.00"

        # ── Otomatik sütun genişliği ──────────────────────────────────────────
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[len(str(r.get(col_name, "") or "")) for r in rows]
            )
            # Geniş sütun başlıkları için ekstra tolerans
            worksheet.column_dimensions[col_letter].width = min(max_len + 4, 35)

        # Satır yükseklikleri
        worksheet.row_dimensions[1].height = 30
        worksheet.row_dimensions[2].height = 25

        # Satırları dondur (başlıklar sabit kalsın)
        worksheet.freeze_panes = "A3"

    buffer.seek(0)
    return buffer.getvalue()
